import openpyxl

wb = openpyxl.load_workbook('题库(622道).xlsx')
ws = wb['Sheet1']

# Check TF question cells for color (col 3 = title, col 4 = full question)
print("=== True/False - Checking question cell FILL colors ===")
tf_count = 0
for row in ws.iter_rows(min_row=4):
    qtype = row[1].value
    if qtype and '判断' in str(qtype):
        tf_count += 1
        title_cell = row[2]  # col 3: 题目 title
        q_cell = row[3]      # col 4: full question
        a_cell = row[4]      # col 5: answer

        # Check background fill
        for label, cell in [('Title', title_cell), ('Q', q_cell), ('Answer', a_cell)]:
            fill = cell.fill
            font = cell.font
            print(f'  Row {row[0].value} {label}: fill.fgColor.rgb={fill.fgColor.rgb}, '
                  f'font.color.rgb={font.color.rgb if font.color else None}, '
                  f'font.color.theme={font.color.theme if font.color else None}, '
                  f'value="{str(cell.value)[:50]}"')
        if tf_count >= 4:
            break

# Let me also check the actual answer values for ALL TF questions to see pattern
print("\n=== All TF answers ===")
answers = set()
for row in ws.iter_rows(min_row=4):
    qtype = row[1].value
    if qtype and '判断' in str(qtype):
        a = row[4].value
        if a:
            answers.add(str(a).strip())
print(f"Unique TF answers: {answers}")
print(f"TF count: {sum(1 for row in ws.iter_rows(min_row=4) if row[1].value and '判断' in str(row[1].value))}")

# Check for any red/colored fonts anywhere in coding questions
print("\n=== Coding Q - checking for any colored formatting ===")
code_count = 0
for row in ws.iter_rows(min_row=4):
    qtype = row[1].value
    if qtype and '编程' in str(qtype):
        code_count += 1
        q_cell = row[3]
        a_cell = row[4]

        # Check if cell has any inline formatting via CellRichText
        from openpyxl.cell.rich_text import CellRichText
        print(f'Row {row[0].value}: Q is RichText={isinstance(q_cell.value, CellRichText)}, '
              f'A is RichText={isinstance(a_cell.value, CellRichText)}')
        if code_count >= 2:
            break

# Check: do any cells in the file use RichText at all?
print("\n=== File-wide RichText check ===")
rt_rows = []
for row in ws.iter_rows(min_row=1):
    for col_idx, cell in enumerate(row):
        from openpyxl.cell.rich_text import CellRichText
        if isinstance(cell.value, CellRichText):
            rt_rows.append((cell.row, col_idx, str(cell.value)[:80]))
print(f"Total RichText cells: {len(rt_rows)}")
for r in rt_rows[:5]:
    print(f"  Row {r[0]}, Col {r[1]}: {r[2]}")
