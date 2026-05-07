import openpyxl
from openpyxl.cell.rich_text import CellRichText

wb = openpyxl.load_workbook('题库(622道).xlsx')
ws = wb['Sheet1']

# Check true/false questions - look at font/color in answer column
tf_count = 0
for row in ws.iter_rows(min_row=4):
    qtype = row[1].value
    if qtype and '判断' in str(qtype):
        tf_count += 1
        answer_cell = row[4]
        font = answer_cell.font
        fill = answer_cell.fill
        print(f'Row: {row[0].value}, Answer: "{answer_cell.value}", Font color: {font.color}, Fill fg: {fill.fgColor}')
        if tf_count >= 6:
            break

print()

# Check coding questions - look for red font parts in answer
code_count = 0
for row in ws.iter_rows(min_row=4):
    qtype = row[1].value
    if qtype and '编程' in str(qtype):
        code_count += 1
        ans_cell = row[4]
        print(f'Row: {row[0].value}, Title: {str(row[2].value)[:60]}')

        if isinstance(ans_cell.value, CellRichText):
            print('  Answer is RichText!')
            for seg in ans_cell.value:
                text_preview = str(seg.text).replace('\n', '\\n')[:100]
                print(f'    Text: "{text_preview}"')
                print(f'    Font: color={seg.font.color}, bold={seg.font.bold}')
        else:
            text_preview = str(ans_cell.value).replace('\n', '\\n')[:200]
            print(f'  Answer (plain): "{text_preview}"')

            # Check cell-level font
            print(f'  Cell font color: {ans_cell.font.color}')

        print()
        if code_count >= 3:
            break

# Also check: are there any rich text cells in the Q column?
print('=== Checking for RichText cells in col 4 (full question) ===')
rt_count = 0
for row in ws.iter_rows(min_row=4):
    if isinstance(row[3].value, CellRichText):
        rt_count += 1
        if rt_count <= 3:
            print(f'Row {row[0].value}: RichText in Q, type={row[1].value}')
print(f'Total RichText Q cells: {rt_count}')
