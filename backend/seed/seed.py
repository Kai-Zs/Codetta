"""
将 题库(622道).xlsx 和 2544名单.xlsx 导入 SQLite 数据库。

Excel 结构: row[0]=注释, row[1]=表头, row[2:]=数据(含章节标题行)
列: 题号 | 题型 | 标题 | 题目 | 答案 | 备注
"""
import json
import os
import re
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl
from app.database import get_db, init_db

ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
EXCEL_QA = os.path.join(ROOT, "data", "题库(622道).xlsx")
EXCEL_ROSTER = os.path.join(ROOT, "data", "2544名单.xlsx")
HTML_PROG = os.path.join(ROOT, "data", "编程题抽出来的题库.files", "sheet001.htm")


def load_qa_rows(path):
    """返回清洗后的题目行，跳过注释行和章节标题行"""
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    if ws.max_row < 3:
        return []
    result = []
    for r in range(3, ws.max_row + 1):
        cell_qn = ws.cell(row=r, column=1)
        if cell_qn.value is None:
            continue
        qn = str(cell_qn.value).strip().lstrip("'")
        if not qn or qn.startswith("第") and "章" in qn:
            continue

        def cell_str(col):
            v = ws.cell(row=r, column=col).value
            return str(v).strip() if v is not None else ""

        result.append({
            "q_number": qn,
            "type": cell_str(2),
            "title": cell_str(3),
            "content": cell_str(4),
            "answer": cell_str(5),
            "note": cell_str(6),
        })
    return result


def load_roster_rows(path):
    """读取名单 Excel，返回 [(student_id, name), ...]"""
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    result = []
    for r in range(2, ws.max_row + 1):
        sid = ws.cell(row=r, column=1).value
        name = ws.cell(row=r, column=2).value
        if sid is None:
            continue
        sid = str(sid).strip()
        if sid.isdigit() and len(sid) == 10:
            result.append((sid, str(name).strip() if name else sid[-4:] + "同学"))
    return result


def clean_dollar(val):
    """清洗内容/标题/备注字段中的 $ 转义残留"""
    if not isinstance(val, str):
        return val
    val = re.sub(r"'?\$'?w\+'?", "'w+'", val)
    val = re.sub(r"'?\$-inf'?", "-inf", val)
    val = re.sub(r"'?\+inf\$'?", "+inf", val)
    val = val.replace("$\n", "\n").replace("\n$", "\n")
    val = re.sub(r"\$\$+", "", val)
    val = re.sub(r"^'?\$'?", "", val)
    val = re.sub(r"'?\$'?$", "", val)
    return val.strip()


def clean_answer_dollar(val):
    """仅处理答案中确定是 Excel 转义的 $（如 $True → True），保留填空分隔符 $"""
    if not isinstance(val, str):
        return val
    val = val.strip()
    val = val.lstrip("$")
    return val


def parse_prog_html(path):
    """解析编程题 HTML 表格：td[4] 中 font10(红)=答案，其余=模板"""
    if not os.path.exists(path):
        return {}
    html = None
    for enc in ("utf-8", "gbk", "gb2312", "gb18030"):
        try:
            with open(path, encoding=enc) as f:
                html = f.read()
            break
        except (UnicodeDecodeError, UnicodeError):
            continue
    if html is None:
        print(f"无法解码 HTML 文件: {path}")
        return {}

    prog_map = {}
    rows = re.findall(r'<tr[^>]*>(.*?)</tr>', html, re.DOTALL)
    for row in rows:
        tds = re.findall(r'<td[^>]*>(.*?)</td>', row, re.DOTALL)
        if len(tds) < 5:
            continue
        qn = re.sub(r'<[^>]+>', '', tds[0]).strip()
        if not re.match(r'^\d+\.\d+', qn):
            continue

        # td[4] = 代码列
        code_html = tds[4]
        template_parts = []
        answer_parts = []
        full_parts = []  # 按原始顺序收集所有片段，用于生成完整代码
        # 按 font10 标签分割（font 和 class 可能跨行）
        parts = re.split(r'(<font[^>]*class="font10"[^>]*>.*?</font>)', code_html, flags=re.DOTALL)
        for part in parts:
            m = re.match(r'<font[^>]*class="font10"[^>]*>(.*?)</font>', part, re.DOTALL)
            if m:
                answer_parts.append(m.group(1))
            else:
                template_parts.append(part)
            full_parts.append(part)

        def clean_html(s):
            # 1. <br> → 占位符；&nbsp; → 空格占位符
            s = re.sub(r'<br\s*/?>', '\x00BR\x00', s)
            s = s.replace('&nbsp;', '\x00SP\x00')
            # 2. 去掉 ruby 注解和所有 HTML 标签
            s = re.sub(r'<ruby>.*?</ruby>', '', s, flags=re.DOTALL)
            s = re.sub(r'<[^>]+>', '', s)
            # 3. 其他 HTML 实体
            s = s.replace('&amp;', '&').replace('&lt;', '<').replace('&gt;', '>').replace('&quot;', '"')
            # 4. 把所有空白序列压缩为单个空格（占位符不受影响）
            s = re.sub(r'\s+', ' ', s)
            # 5. 恢复占位符
            s = s.replace('\x00BR\x00', '\n')
            s = s.replace('\x00SP\x00', ' ')
            # 6. 每行仅去行尾空白，保留缩进
            lines = [l.rstrip() for l in s.split('\n')]
            lines = [l for l in lines if l.strip()]
            # 7. 合并单独成行的 # 与下一行
            merged = []
            i = 0
            while i < len(lines):
                if lines[i] == '#' and i + 1 < len(lines):
                    merged.append('# ' + lines[i+1])
                    i += 2
                else:
                    merged.append(lines[i])
                    i += 1
            return '\n'.join(merged)

        prog_map[qn] = {
            "template": clean_html("".join(template_parts)),
            "answer_code": clean_html("".join(answer_parts)),
            "full_code": clean_html("".join(full_parts)),
        }
    return prog_map


def parse_single_choice_options(content):
    """从题干文本中提取 A/B/C/D 选项，选项可能跨行"""
    # 用 A./B./C./D. 或 A、/B、/C、/D、 作为分隔标记
    pattern = r'([A-D])[.、．]\s*(.+?)(?=\n?[A-D][.、．]|\Z)'
    matches = re.findall(pattern, content, re.DOTALL)
    if matches:
        return [f"{m[0]}. {m[1].strip()}" for m in matches]
    return []


def match_answer_to_option(raw_answer, options):
    """去除选项的 A. 前缀后与答案精确比对，返回匹配的字母，否则 None"""
    ans_text = raw_answer.strip().lstrip("$")
    for opt_text in options:
        # 取选项文本（去掉 "A. " 前缀）
        m = re.match(r'[A-D][.、．]\s*(.+)', opt_text, re.DOTALL)
        if m:
            opt_body = m.group(1).strip()
            if ans_text == opt_body:
                return opt_text[0]
    return None


def seed_questions(db):
    prog_map = parse_prog_html(HTML_PROG)
    rows = load_qa_rows(EXCEL_QA)
    skipped = []

    for row in rows:
        q_number = row["q_number"]
        qtype = row["type"]
        title = clean_dollar(row["title"])
        content = clean_dollar(row["content"])
        raw_answer = row["answer"]
        note = clean_dollar(row["note"])
        chapter = q_number.split(".")[0] if "." in q_number else q_number
        options = None
        answer_parts = None
        template = None
        answer_code = None
        is_active = 1

        if qtype == "单选题":
            raw_answer = clean_answer_dollar(raw_answer)
            options = parse_single_choice_options(content)
            if options:
                answer_letter = match_answer_to_option(raw_answer, options)
                if answer_letter:
                    raw_answer = answer_letter
                else:
                    skipped.append(f"{q_number}: 单选答案匹配失败，答案=[{raw_answer[:80]}]，需手动处理")
            else:
                skipped.append(f"{q_number}: 单选未解析到选项，需手动处理")

        elif qtype == "判断题":
            raw_answer = raw_answer.strip()
            if raw_answer not in ("正确", "错误"):
                skipped.append(f"{q_number}: 判断题答案异常 [{raw_answer[:50]}]，需手动处理")

        elif qtype == "填空题":
            answer_parts_list = [p.strip() for p in raw_answer.split("$") if p.strip()]
            if not answer_parts_list:
                answer_parts_list = [raw_answer.strip()]
            answer_parts = json.dumps(answer_parts_list, ensure_ascii=False)

        elif qtype == "编程题":
            prog_info = prog_map.get(q_number, {})
            template = prog_info.get("template", "")
            answer_code = prog_info.get("full_code", "")  # 完整代码（模板+答案）
            if not raw_answer.strip() and not answer_code:
                is_active = 0
                skipped.append(f"{q_number}: 编程题无答案，标记为停用")

        if q_number in ("8.33", "8.34", "8.35", "8.36"):
            is_active = 0

        db.execute("""
            INSERT OR REPLACE INTO questions (q_number, chapter, type, title, content, options,
                                              answer, answer_parts, template, answer_code, note, is_active)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (q_number, chapter, qtype, title, content,
              json.dumps(options, ensure_ascii=False) if options else None,
              raw_answer,
              answer_parts,
              template, answer_code, note, is_active))

    return skipped


def seed_roster(db):
    entries = load_roster_rows(EXCEL_ROSTER)
    for sid, name in entries:
        db.execute("""
            INSERT OR IGNORE INTO users (student_id, name, in_roster)
            VALUES (?, ?, 1)
        """, (sid, name))


def main():
    try:
        init_db()
    except Exception as e:
        print(f"init_db 出错: {e}")
        return
    with get_db() as db:
        db.execute("DELETE FROM progress")
        db.execute("DELETE FROM questions")
        db.execute("DELETE FROM users")
        skipped = seed_questions(db)
        seed_roster(db)
    print(f"种子数据导入完成。")
    if skipped:
        print(f"\n警告 {len(skipped)} 条：")
        for s in skipped:
            print(f"  - {s}")


if __name__ == "__main__":
    main()
